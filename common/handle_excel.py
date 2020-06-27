#!D:/PychramCode
# -*- coding: utf-8 -*-
# @Time : 2020/6/19 15:30
# @Author : 涛涛
# @Software: PyCharm

import openpyxl


class HandleExcel:
    """
    处理excel数据
    """

    def __init__(self, filename, sheetname):
        """
        初始化对象属性
        :param filename:excel路径
        :param sheetname:表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    # 读excel
    def read_excel(self):
        # 获取工作薄对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择指定表单
        sh = wb[self.sheetname]
        # 按行获取所有数据，转换为，列表
        rows_data = list(sh.rows)
        # 选择表头放入到列表中
        title = []
        for i in rows_data[0]:
            title.append(i.value)
        # 创建一个空列表，用来保存所有的用例数据
        case_datas = []

        # 遍历除表头外的每一行数据
        for item in rows_data[1:]:
            # 每遍历一行数据，就创建一个空列表，来存放该行数据
            values = []
            for i in item:
                values.append(i.value)
            # 将该行数据和表头进行打包处理，转换为字典
            case = dict(zip(title, values))
            # 将打包后的字典，放入case_datas
            case_datas.append(case)
        # 返回读取出来的所有数据
        return case_datas

    def write_excel(self, row, column, value):
        """
        写入数据
        :param row: 行
        :param column: 列
        :param value: 要写入的值
        :return:
        """
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheetname]
        # 根据行、列去写入内容
        sh.cell(row = row,column = column,value = value)
        # 把工作簿保存为文件
        wb.save(self.filename)

if __name__ == '__main__':
    wb = HandleExcel(r"D:\PychramCode\auto_api_test\data\apicases.xlsx","register")
    wb.write_excel(8,8,"111")